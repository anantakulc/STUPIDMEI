"""Render a 10-tab institutional Excel workbook from the narrative + valuation JSONs.

Same JSON in -> same xlsx out. The agent does not touch Excel layout.

Usage:
    python render_excel.py --ticker AVGO --category AI
       reads:  output/AVGO/AVGO.json + output/AVGO/AVGO_valuation.json
       writes: output/AVGO/AVGO.xlsx

Tabs (locked order):
    1. Cover                — institutional tear sheet (rating, PT, key metrics)
    2. Market Stats         — current snapshot (price, mkt cap, EV, ADTV, vol, beta)
    3. Historicals          — multi-year P&L table
    4. Assumptions          — every input + per-component reasoning
    5. WACC & DCF           — WACC adjudication + full FCFF build + dual TV
    6. Formula trace        — step-by-step derivation of every number
    7. Sensitivity          — WACC × g (DCF) or Ke × g (DDM) matrix
    8. Relative & SOTP      — peer multiples + multi-method relative + segment SOTP
    9. Scenarios            — DCF scenarios + 12m price scenarios + prob-weighted target
    10. Reasoning log       — methodology narrative
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")


# ============================================================
# STYLE PRIMITIVES
# ============================================================

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
H2_FONT = Font(name="Calibri", size=12, bold=True, color="111827")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="6B7280")
BODY_FONT = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)
MUTED_FONT = Font(name="Calibri", size=9, color="6B7280")
MONO_FONT = Font(name="Consolas", size=10)

HEADER_FILL = PatternFill("solid", fgColor="111827")
SECTION_FILL = PatternFill("solid", fgColor="F3F4F6")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="DBEAFE")
GOOD_FILL = PatternFill("solid", fgColor="DCFCE7")
BAD_FILL = PatternFill("solid", fgColor="FEE2E2")
NEUTRAL_FILL = PatternFill("solid", fgColor="FEF3C7")
WARN_FILL = PatternFill("solid", fgColor="FED7AA")

THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ============================================================
# UTILITIES
# ============================================================

def fmt_currency(v, currency):
    """Short currency formatter for tear sheets."""
    if v is None:
        return "n/a"
    if currency == "IDR":
        return f"IDR {v:,.0f}"
    if abs(v) >= 100:
        return f"${v:,.0f}"
    return f"${v:,.2f}"


def fmt_bn(v, currency="USD"):
    """Billions formatter."""
    if v is None:
        return "n/a"
    if currency == "IDR":
        return f"IDR {v:,.1f}T"  # for IDR, use trillions
    return f"${v:,.1f}B"


def fmt_pct(v, decimals=1):
    if v is None:
        return "n/a"
    return f"{v*100:.{decimals}f}%"


def autosize(ws, max_width=80):
    """Compute column widths from cell contents."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    cell_str = str(cell.value)
                    if len(cell_str) > max_len:
                        max_len = len(cell_str)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def section_header(ws, row, label, span=6):
    """Write a section divider row."""
    cell = ws.cell(row=row, column=1, value=label.upper())
    cell.font = H2_FONT
    cell.fill = SECTION_FILL
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def kv_row(ws, row, kv_pairs):
    """Write a row of key-value pairs in a 2-col grid pattern (label, value, label, value...)."""
    col = 1
    for label, value in kv_pairs:
        ws.cell(row=row, column=col, value=label).font = BODY_BOLD
        ws.cell(row=row, column=col + 1, value=value).font = BODY_FONT
        col += 2


# ============================================================
# 1. COVER (institutional tear sheet)
# ============================================================

def build_cover(wb, narrative, valuation):
    ws = wb.create_sheet("Cover")
    ticker = narrative["ticker"]
    name = narrative["name"]
    rec = narrative.get("recommendation", {})
    currency = valuation.get("currency", "USD")
    primary = valuation.get("primary_method", {})
    primary_outputs = primary.get("outputs", {})

    ws["A1"] = f"{name.upper()} ({ticker}) — Valuation Model"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Institutional Equity Research  |  Date: {narrative.get('date', '')}  |  Listings: {narrative.get('listings', '')}  |  Engine: himself65/finance-skills"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    snapshot = {s["label"]: s["value"] for s in narrative.get("snapshot", []) if isinstance(s, dict)}

    def snap(key, default="n/a"):
        for k, v in snapshot.items():
            if k.lower().startswith(key.lower()):
                return v
        return default

    rec_action = rec.get("action", "")
    rec_color_fill = None
    if rec.get("tone") == "positive":
        rec_color_fill = GOOD_FILL
    elif rec.get("tone") == "negative":
        rec_color_fill = BAD_FILL
    else:
        rec_color_fill = NEUTRAL_FILL

    rows = [
        ("Rating", rec_action, "Price target", fmt_currency(rec.get("target_12m"), currency)),
        ("Current price", fmt_currency(rec.get("current_price"), currency),
         "Upside to target", f"{rec.get('upside_pct', 0):+.1f}%"),
        ("Blended (math)", fmt_currency(valuation.get("blended_target"), currency),
         "Math upside vs spot", f"{valuation.get('upside_pct', 0):+.1f}%"),
        ("Market cap", snap("market cap"), "Enterprise value", snap("enterprise")),
        ("FY revenue", snap("FY") or snap("revenue"), "FY EBITDA / margin", snap("EBITDA")),
        ("FY FCF", snap("FCF"), "Net debt", snap("Net debt")),
        ("Forward EPS", snap("Forward EPS"), "Forward P/E", snap("Forward P/E")),
        ("EV/EBITDA (TTM)", snap("EV"), "Beta", snap("Beta")),
        ("ADTV (30d)", snap("ADTV"), "30d vol", snap("30d")),
        ("Analyst consensus", snap("consensus"), "Mean PT", snap("Mean") or snap("median")),
    ]

    # WACC / Ke row depends on method
    if primary.get("name") == "DCF":
        wacc_value = primary.get("inputs", {}).get("wacc", {}).get("value")
        ke_value = None
        comps = primary.get("inputs", {}).get("wacc", {}).get("components", {})
        if comps:
            ke_value = comps.get("rf", 0) + comps.get("beta", 0) * comps.get("erp", 0)
        rows.append(("WACC", fmt_pct(wacc_value, 2), "Ke (CAPM)", fmt_pct(ke_value, 2)))
    elif primary.get("name") == "DDM":
        ke_value = primary.get("outputs", {}).get("cost_of_equity_pct")
        rows.append(("Cost of equity (Ke)", f"{ke_value:.2f}%" if ke_value is not None else "n/a",
                     "Method", "DDM"))

    # Implied prices summary
    if primary.get("name") == "DCF":
        rows.append(("DCF Base implied", fmt_currency(primary_outputs.get("implied_px"), currency),
                     "Cross-check implied",
                     fmt_currency(valuation.get("cross_check", {}).get("outputs", {}).get("implied_px") if valuation.get("cross_check") else None, currency)))
    elif primary.get("name") == "DDM":
        ddm_val = primary_outputs.get("ddm", {}).get("implied_value")
        jpb_val = primary_outputs.get("justified_pb", {}).get("implied_value")
        rows.append(("DDM implied", fmt_currency(ddm_val, currency),
                     "Justified P/B implied", fmt_currency(jpb_val, currency)))

    row = 4
    for r in rows:
        ws.cell(row=row, column=1, value=r[0]).font = BODY_BOLD
        ws.cell(row=row, column=2, value=r[1]).font = BODY_FONT
        ws.cell(row=row, column=3, value=r[2]).font = BODY_BOLD
        ws.cell(row=row, column=4, value=r[3]).font = BODY_FONT
        if r[0] == "Rating":
            ws.cell(row=row, column=2).fill = rec_color_fill
            ws.cell(row=row, column=2).font = BODY_BOLD
        if r[0] == "Upside to target" or r[2] == "Upside to target":
            # color the upside cell
            up = rec.get("upside_pct", 0)
            ws.cell(row=row, column=4).fill = GOOD_FILL if up >= 0 else BAD_FILL
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Three-line thesis").font = H2_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for i, t in enumerate(narrative.get("thesis", []), start=1):
        ws.cell(row=row, column=1, value=f"{i}.").font = BODY_BOLD
        cell = ws.cell(row=row, column=2, value=t)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 45
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Research cycle").font = MUTED_FONT
    ws.cell(row=row, column=2, value="equity-bull-case skill → equity-bear-case skill (independent) → synthesis → DCF/DDM compute → render Excel").font = MUTED_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value="Disclaimer").font = MUTED_FONT
    ws.cell(row=row, column=2, value="For research only. Not financial advice. Engine: himself65/finance-skills + custom DCF/DDM compute scripts.").font = MUTED_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

    autosize(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18


# ============================================================
# 2. MARKET STATS
# ============================================================

def build_market_stats(wb, narrative, valuation):
    ws = wb.create_sheet("Market Stats")
    ws["A1"] = f"{narrative['ticker']} — Market stats (snapshot)"
    ws["A1"].font = TITLE_FONT

    headers = ["Metric", "Value", "Notes"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row = 4
    for s in narrative.get("snapshot", []):
        if isinstance(s, dict):
            ws.cell(row=row, column=1, value=s.get("label", "")).font = BODY_FONT
            ws.cell(row=row, column=2, value=s.get("value", "")).font = BODY_BOLD
            row += 1

    autosize(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50


# ============================================================
# 3. HISTORICALS
# ============================================================

def build_historicals(wb, narrative):
    hist = narrative.get("historical_financials") or []
    if not hist:
        return
    ws = wb.create_sheet("Historicals")
    ws["A1"] = f"{narrative['ticker']} — Historical financials (USD billions unless noted)"
    ws["A1"].font = TITLE_FONT

    # Pivot: rows = metrics, cols = fiscal years
    metrics = [
        ("Revenue ($B)", "revenue_usd_b"),
        ("Revenue growth YoY", "revenue_growth_yoy_pct"),
        ("Gross margin", "gross_margin_pct"),
        ("EBITDA ($M)", "ebitda_usd_m"),
        ("EBITDA margin", "ebitda_margin_pct"),
        ("Net income ($M)", "net_income_usd_m"),
        ("FCF ($M)", "fcf_usd_m"),
        ("EPS", "eps"),
    ]
    # Header row
    ws.cell(row=3, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    for c, h in enumerate(hist, start=2):
        cell = ws.cell(row=3, column=c, value=h.get("fy", ""))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for r, (label, key) in enumerate(metrics, start=4):
        ws.cell(row=r, column=1, value=label).font = BODY_BOLD
        for c, h in enumerate(hist, start=2):
            v = h.get(key)
            if v is None:
                ws.cell(row=r, column=c, value="n/a").font = MUTED_FONT
            else:
                if key.endswith("_pct"):
                    ws.cell(row=r, column=c, value=v / 100).number_format = "0.0%"
                else:
                    ws.cell(row=r, column=c, value=round(v, 2) if isinstance(v, float) else v)
    autosize(ws)
    ws.column_dimensions["A"].width = 24


# ============================================================
# 4. ASSUMPTIONS (per-component reasoning)
# ============================================================

def build_assumptions(wb, valuation):
    ws = wb.create_sheet("Assumptions")
    primary = valuation.get("primary_method", {})
    method = primary.get("name", "")

    ws["A1"] = f"Assumptions — {method}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "EDIT cells in column B to flex the model. Source in column C. Reasoning in column D."
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:D2")

    headers = ["Input", "Value", "Source", "Reasoning"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    inputs = primary.get("inputs", {}) or {}
    row = 5

    def write_input(label, value, source="", reasoning=""):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = BODY_BOLD
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        ws.cell(row=row, column=3, value=source).font = MUTED_FONT
        ws.cell(row=row, column=4, value=reasoning).font = MUTED_FONT
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    if method == "DCF":
        wacc = inputs.get("wacc", {})
        write_input("WACC", wacc.get("value"), "computed", wacc.get("reasoning", ""))
        components_reasoned = wacc.get("components_reasoned") or {}
        if components_reasoned:
            for k, comp in components_reasoned.items():
                v = comp.get("value") if isinstance(comp, dict) else comp
                r = comp.get("reasoning", "") if isinstance(comp, dict) else ""
                write_input(f"  - {k}", v, "input", r)
        else:
            for k, v in (wacc.get("components") or {}).items():
                write_input(f"  - {k}", v, "input", "")

        tg = inputs.get("terminal_growth", {})
        write_input("Terminal growth", tg.get("value"), "input", tg.get("reasoning", ""))
        write_input("Forecast horizon (yrs)", inputs.get("forecast_horizon_years"), "input", "")

        row += 1
        section_header(ws, row, "FCF projections", span=4)
        row += 1
        ws.cell(row=row, column=1, value="Year").font = BODY_BOLD
        ws.cell(row=row, column=2, value="FCF (bn)").font = BODY_BOLD
        ws.cell(row=row, column=3, value="Revenue (bn)").font = BODY_BOLD
        ws.cell(row=row, column=4, value="Rationale").font = BODY_BOLD
        row += 1
        for proj in inputs.get("fcf_projections", []):
            ws.cell(row=row, column=1, value=proj.get("year"))
            ws.cell(row=row, column=2, value=proj.get("fcf_b"))
            ws.cell(row=row, column=3, value=proj.get("revenue_b", "n/a"))
            ws.cell(row=row, column=4, value=proj.get("rationale", "")).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    elif method == "DDM":
        for label, key, reasoning_field in [
            ("Risk-free rate (rf)", "rf", "rf_reasoning"),
            ("Equity risk premium (ERP)", "erp", "erp_reasoning"),
            ("Beta", "beta", "beta_reasoning"),
            ("Sustainable D0", "d0", "d0_reasoning"),
            ("Book value / share", "book_value_per_share", "book_value_reasoning"),
            ("ROE", "roe", "roe_reasoning"),
            ("g_high (years 1-5)", "g_high", "g_high_reasoning"),
            ("g_terminal", "g_terminal", "g_terminal_reasoning"),
            ("g_book", "g_book", "g_book_reasoning"),
            ("Years high growth", "years_high", None),
            ("Years declining", "years_decline", None),
            ("Excess returns horizon", "horizon_excess_returns", None),
            ("Max sustainable payout", "max_payout_cap", None),
        ]:
            reasoning_text = inputs.get(reasoning_field, "") if reasoning_field else ""
            write_input(label, inputs.get(key), "input/default", reasoning_text)

    elif method == "SOTP":
        section_header(ws, row, "Segments", span=4)
        row += 1
        for seg in inputs.get("segments", []) or primary.get("outputs", {}).get("segments", []):
            write_input(seg["name"], seg.get("method", ""), "input", seg.get("reasoning", ""))

    autosize(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60


# ============================================================
# 5. WACC & DCF (full FCFF build + adjudication + dual TV)
# ============================================================

def build_wacc_dcf(wb, valuation, narrative):
    primary = valuation.get("primary_method", {})
    method = primary.get("name", "")
    if method not in ("DCF", "DDM", "SOTP"):
        return
    ws = wb.create_sheet("WACC & DCF" if method == "DCF" else f"{method} build")
    currency = valuation.get("currency", "USD")
    outputs = primary.get("outputs", {})
    inputs = primary.get("inputs", {})

    ws["A1"] = f"{narrative['ticker']} — {method} build"
    ws["A1"].font = TITLE_FONT

    row = 3

    if method == "DCF":
        # === WACC inputs & adjudication ===
        section_header(ws, row, "WACC inputs & adjudication", span=7)
        row += 1
        ws.cell(row=row, column=1, value="Input").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value="Source").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=3, value="Value").font = HEADER_FONT
        ws.cell(row=row, column=3).fill = HEADER_FILL
        row += 1

        comps = inputs.get("wacc", {}).get("components", {})
        rf = comps.get("rf", 0)
        beta = comps.get("beta", 0)
        erp = comps.get("erp", 0)
        debt_w = comps.get("debt_weight", 0)
        kd_at = comps.get("cost_of_debt_after_tax", 0)
        ke = rf + beta * erp

        def write_wacc_input(label, source, value, value_pct=False):
            nonlocal row
            ws.cell(row=row, column=1, value=label).font = BODY_BOLD
            ws.cell(row=row, column=2, value=source).font = MUTED_FONT
            cell = ws.cell(row=row, column=3, value=value)
            if value_pct:
                cell.number_format = "0.00%"
            cell.font = BODY_FONT
            row += 1

        write_wacc_input("Risk-free rate (rf)", "10Y government bond yield", rf, True)
        write_wacc_input("Beta", "input.json or yfinance sanitized", beta)
        write_wacc_input("Equity risk premium (ERP)", "Damodaran / DM-EM standard", erp, True)
        write_wacc_input("ke = rf + β × ERP", "CAPM", ke, True)
        write_wacc_input("Debt weight (D/V)", "input.json", debt_w, True)
        write_wacc_input("Equity weight (E/V)", "1 − D/V", 1 - debt_w, True)
        write_wacc_input("kd (after-tax)", "input.json", kd_at, True)
        wacc_value = inputs.get("wacc", {}).get("value", 0)
        write_wacc_input("Formula WACC", "(E/V × ke) + (D/V × kd)", wacc_value, True)

        # === Adjudication crosscheck ===
        adj = outputs.get("wacc_adjudication") or {}
        if adj:
            row += 1
            section_header(ws, row, "Adjudication crosscheck", span=7)
            row += 1
            band = adj.get("sector_band", [0, 0])
            ws.cell(row=row, column=1, value=f"Sector band ({adj.get('sector', 'default')})").font = BODY_BOLD
            ws.cell(row=row, column=3, value=f"{band[0]*100:.1f}% − {band[1]*100:.1f}%").font = BODY_FONT
            row += 1
            ws.cell(row=row, column=1, value="Formula WACC vs band").font = BODY_BOLD
            verdict = adj.get("verdict", "")
            verdict_cell = ws.cell(row=row, column=3, value=f"{wacc_value*100:.2f}% — {verdict}")
            if "BELOW" in verdict.upper() or "ABOVE" in verdict.upper():
                verdict_cell.fill = WARN_FILL
            else:
                verdict_cell.fill = GOOD_FILL
            verdict_cell.font = BODY_BOLD
            row += 1
            if adj.get("notes"):
                ws.cell(row=row, column=1, value="Notes").font = BODY_BOLD
                ws.cell(row=row, column=3, value=adj["notes"]).font = MUTED_FONT
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
                row += 1

        # === FCFF build ===
        row += 1
        section_header(ws, row, "DCF — FCFF projection", span=8)
        row += 1
        fcff_build = outputs.get("fcff_build") or []
        years = [f.get("year", "") for f in fcff_build]
        # If no rich build, fall back to FCF projection only
        if any(f.get("has_full_build") for f in fcff_build):
            # Full FCFF build available
            ws.cell(row=row, column=1, value="Item").font = HEADER_FONT
            ws.cell(row=row, column=1).fill = HEADER_FILL
            for c, y in enumerate(years, start=2):
                cell = ws.cell(row=row, column=c, value=str(y))
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            row += 1
            metric_rows = [
                ("Revenue (bn)", "revenue_b", None),
                ("YoY growth", "revenue_growth_pct", "pct"),
                ("EBIT margin", "ebit_margin", "pct"),
                ("EBIT (bn)", "ebit_b", None),
                ("Tax rate", "tax_rate", "pct"),
                ("NOPAT (bn)", "nopat_b", None),
                ("D&A (bn)", "da_b", None),
                ("Capex (bn)", "capex_b", None),
                ("ΔNWC (bn)", "wc_change_b", None),
                ("FCFF (bn)", "fcf_b", None),
            ]
            for label, key, fmt in metric_rows:
                ws.cell(row=row, column=1, value=label).font = BODY_BOLD
                for c, fb in enumerate(fcff_build, start=2):
                    # Some keys map differently - revenue_growth_pct is on input projection, not fcff_build
                    inputs_proj = inputs.get("fcf_projections", [])
                    if key == "revenue_growth_pct" and c - 2 < len(inputs_proj):
                        v = inputs_proj[c - 2].get("revenue_growth_pct")
                        if v is not None:
                            v = v / 100  # they store as e.g. 65.9 meaning 65.9%
                    elif key == "ebit_margin":
                        v = fb.get("ebit_margin")
                    elif key == "tax_rate":
                        v = fb.get("tax_rate")
                    else:
                        v = fb.get(key)
                    if v is None:
                        ws.cell(row=row, column=c, value="n/a").font = MUTED_FONT
                    elif fmt == "pct":
                        cell = ws.cell(row=row, column=c, value=v)
                        cell.number_format = "0.0%"
                    else:
                        ws.cell(row=row, column=c, value=round(v, 3) if isinstance(v, float) else v)
                row += 1
            # PV row
            trace = outputs.get("calculation_trace", {})
            per_year = trace.get("explicit_fcf", {}).get("per_year", [])
            if per_year:
                ws.cell(row=row, column=1, value="Discount factor").font = BODY_BOLD
                for c, py in enumerate(per_year, start=2):
                    cell = ws.cell(row=row, column=c, value=py.get("discount_factor"))
                    cell.number_format = "0.0000"
                row += 1
                ws.cell(row=row, column=1, value="PV of FCFF (bn)").font = BODY_BOLD
                for c, py in enumerate(per_year, start=2):
                    cell = ws.cell(row=row, column=c, value=py.get("pv_b"))
                    cell.fill = HIGHLIGHT_FILL
                row += 1
        else:
            # Lump-sum FCF only (legacy / minimal inputs)
            ws.cell(row=row, column=1, value="Item").font = HEADER_FONT
            ws.cell(row=row, column=1).fill = HEADER_FILL
            for c, y in enumerate(years, start=2):
                cell = ws.cell(row=row, column=c, value=str(y))
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            row += 1
            ws.cell(row=row, column=1, value="FCF (bn)").font = BODY_BOLD
            for c, fb in enumerate(fcff_build, start=2):
                ws.cell(row=row, column=c, value=fb.get("fcf_b"))
            row += 1
            trace = outputs.get("calculation_trace", {})
            per_year = trace.get("explicit_fcf", {}).get("per_year", [])
            if per_year:
                ws.cell(row=row, column=1, value="PV (bn)").font = BODY_BOLD
                for c, py in enumerate(per_year, start=2):
                    cell = ws.cell(row=row, column=c, value=py.get("pv_b"))
                    cell.fill = HIGHLIGHT_FILL
                row += 1
            row += 1
            ws.cell(row=row, column=1, value="Note: per-year FCFF decomposition (Revenue/EBIT/NOPAT/D&A/Capex/ΔNWC) not in inputs.json. Add these fields to fcf_projections for full build.").font = MUTED_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        # === Terminal value & EV bridge ===
        row += 1
        section_header(ws, row, "Terminal value & EV bridge", span=4)
        row += 1
        dual_tv = outputs.get("dual_terminal_value") or {}
        ws.cell(row=row, column=1, value="Terminal growth rate (g)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=inputs.get("terminal_growth", {}).get("value")).number_format = "0.00%"
        row += 1
        if fcff_build:
            ws.cell(row=row, column=1, value=f"Terminal FCFF (final year)").font = BODY_BOLD
            ws.cell(row=row, column=2, value=fcff_build[-1].get("fcf_b"))
            row += 1
        ws.cell(row=row, column=1, value="TV — Gordon Growth (bn)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=dual_tv.get("gordon_tv_b"))
        row += 1
        if dual_tv.get("exit_multiple_tv_b") is not None:
            ws.cell(row=row, column=1, value=f"TV — Exit multiple ({dual_tv.get('exit_multiple_x')}× EBITDA, bn)").font = BODY_BOLD
            ws.cell(row=row, column=2, value=dual_tv.get("exit_multiple_tv_b"))
            row += 1
            ws.cell(row=row, column=1, value="TV — Blended (50/50, bn)").font = BODY_BOLD
            ws.cell(row=row, column=2, value=dual_tv.get("blended_tv_b"))
            row += 1
        ws.cell(row=row, column=1, value="PV of TV — Gordon (bn)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=dual_tv.get("gordon_pv_b"))
        row += 1
        if dual_tv.get("blended_pv_b") is not None:
            ws.cell(row=row, column=1, value="PV of TV — Blended (bn)").font = BODY_BOLD
            ws.cell(row=row, column=2, value=dual_tv.get("blended_pv_b"))
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Sum of PV of explicit FCFF (bn)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs.get("pv_explicit_fcf_b"))
        row += 1
        ws.cell(row=row, column=1, value="Implied EV (bn)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs.get("implied_ev_b"))
        ws.cell(row=row, column=2).fill = HIGHLIGHT_FILL
        row += 1
        ws.cell(row=row, column=1, value="(less) Net debt (bn)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=valuation.get("net_debt_b"))
        row += 1
        ws.cell(row=row, column=1, value="Implied equity value (bn)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs.get("implied_equity_b"))
        row += 1
        ws.cell(row=row, column=1, value="Diluted shares (bn)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=valuation.get("shares_outstanding_b"))
        row += 1
        row += 1
        ws.cell(row=row, column=1, value="Implied price per share").font = Font(name="Calibri", size=12, bold=True)
        px_cell = ws.cell(row=row, column=2, value=outputs.get("implied_px"))
        px_cell.fill = HIGHLIGHT_FILL
        px_cell.font = Font(name="Calibri", size=12, bold=True)

    elif method == "DDM":
        # Keep the DDM build similar to before but slightly cleaner
        outputs_ddm = outputs.get("ddm", {})
        outputs_er = outputs.get("excess_returns", {})
        outputs_jpb = outputs.get("justified_pb", {})

        section_header(ws, row, "DDM build", span=3)
        row += 1
        ws.cell(row=row, column=1, value="Year").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value=f"Dividend ({currency})").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        row += 1
        for i, d in enumerate(outputs_ddm.get("explicit_dividends", []), start=1):
            ws.cell(row=row, column=1, value=f"Year {i}")
            ws.cell(row=row, column=2, value=round(d, 2))
            row += 1
        row += 1
        ws.cell(row=row, column=1, value="PV of explicit dividends").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs_ddm.get("pv_explicit"))
        row += 1
        ws.cell(row=row, column=1, value="PV of terminal value").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs_ddm.get("pv_terminal"))
        row += 1
        ws.cell(row=row, column=1, value="DDM implied value").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs_ddm.get("implied_value")).fill = HIGHLIGHT_FILL

        row += 2
        section_header(ws, row, "Excess Returns Model", span=3)
        row += 1
        ws.cell(row=row, column=1, value="Book value (start)").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs_er.get("book_value"))
        row += 1
        ws.cell(row=row, column=1, value="PV of excess returns").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs_er.get("pv_excess"))
        row += 1
        ws.cell(row=row, column=1, value="ER implied value").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs_er.get("implied_value")).fill = HIGHLIGHT_FILL

        row += 2
        section_header(ws, row, "Justified P/B", span=3)
        row += 1
        ws.cell(row=row, column=1, value="Justified P/B ratio").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs_jpb.get("ratio"))
        row += 1
        ws.cell(row=row, column=1, value="JPB implied value").font = BODY_BOLD
        ws.cell(row=row, column=2, value=outputs_jpb.get("implied_value")).fill = HIGHLIGHT_FILL

        row += 2
        ws.cell(row=row, column=1, value="Blended implied price (40/40/20)").font = Font(name="Calibri", size=12, bold=True)
        ws.cell(row=row, column=2, value=outputs.get("implied_px")).fill = HIGHLIGHT_FILL

    autosize(ws)
    if method == "DCF":
        ws.column_dimensions["A"].width = 35
        for c in range(2, 9):
            ws.column_dimensions[get_column_letter(c)].width = 12


# ============================================================
# 6. FORMULA TRACE
# ============================================================

def build_formula_trace(wb, valuation):
    primary = valuation.get("primary_method", {})
    trace = primary.get("calculation_trace") or primary.get("outputs", {}).get("calculation_trace")
    if not trace:
        return
    ws = wb.create_sheet("Formula trace")
    ws["A1"] = "Formula trace"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Step-by-step derivation. Same inputs always produce these numbers."
    ws["A2"].font = SUBTITLE_FONT

    row = 4
    SECTION_ORDER = [
        ("wacc", "WACC / Cost of equity"),
        ("ke", "Cost of equity (Ke)"),
        ("explicit_fcf", "PV of explicit FCF"),
        ("ddm", "DDM value"),
        ("excess_returns", "Excess Returns Model"),
        ("justified_pb", "Justified P/B"),
        ("terminal_value", "Terminal value"),
        ("bridge_to_implied_px", "Bridge to implied price per share"),
        ("blended", "Blended target"),
    ]
    for key, label in SECTION_ORDER:
        calc = trace.get(key)
        if not calc:
            continue
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = BODY_BOLD
        cell.fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        formula_cell = ws.cell(row=row, column=1, value=f"Formula: {calc.get('formula', '')}")
        formula_cell.font = Font(name="Calibri", size=10, italic=True, color="6B7280")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        ws.cell(row=row, column=1, value="Step").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value="Expression").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=3, value="Result").font = HEADER_FONT
        ws.cell(row=row, column=3).fill = HEADER_FILL
        row += 1
        for s in calc.get("steps") or []:
            ws.cell(row=row, column=1, value=s.get("label", "")).font = BODY_FONT
            ws.cell(row=row, column=2, value=s.get("expression", "")).font = MONO_FONT
            ws.cell(row=row, column=3, value=s.get("result", "")).font = BODY_BOLD
            row += 1
        if calc.get("per_year"):
            ws.cell(row=row, column=1, value="Year").font = HEADER_FONT
            ws.cell(row=row, column=1).fill = HEADER_FILL
            ws.cell(row=row, column=2, value="FCF (bn)").font = HEADER_FONT
            ws.cell(row=row, column=2).fill = HEADER_FILL
            ws.cell(row=row, column=3, value="Discount factor").font = HEADER_FONT
            ws.cell(row=row, column=3).fill = HEADER_FILL
            ws.cell(row=row, column=4, value="PV (bn)").font = HEADER_FONT
            ws.cell(row=row, column=4).fill = HEADER_FILL
            row += 1
            for y in calc["per_year"]:
                ws.cell(row=row, column=1, value=y.get("year"))
                ws.cell(row=row, column=2, value=y.get("fcf_b"))
                ws.cell(row=row, column=3, value=y.get("discount_factor"))
                ws.cell(row=row, column=4, value=y.get("pv_b"))
                row += 1
            ws.cell(row=row, column=1, value="Sum (PV explicit FCF)").font = BODY_BOLD
            ws.cell(row=row, column=4, value=calc.get("result")).font = BODY_BOLD
            ws.cell(row=row, column=4).fill = HIGHLIGHT_FILL
            row += 1
        elif calc.get("result"):
            ws.cell(row=row, column=1, value="Result").font = BODY_BOLD
            ws.cell(row=row, column=3, value=calc.get("result")).font = BODY_BOLD
            ws.cell(row=row, column=3).fill = HIGHLIGHT_FILL
            row += 1
        row += 1
    autosize(ws)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14


# ============================================================
# 7. SENSITIVITY
# ============================================================

def build_sensitivity(wb, valuation):
    primary = valuation.get("primary_method", {})
    sens = primary.get("sensitivity")
    if not sens:
        return
    ws = wb.create_sheet("Sensitivity")
    ws["A1"] = f"Sensitivity: {sens.get('rows_label', '')} × {sens.get('cols_label', '')}"
    ws["A1"].font = TITLE_FONT

    rows_label = sens.get("rows_label", "")
    cols_label = sens.get("cols_label", "")
    row_vals = sens.get("row_values", [])
    col_vals = sens.get("col_values", [])
    matrix = sens.get("implied_px_matrix", [])

    ws.cell(row=3, column=1, value=f"{rows_label} \\ {cols_label}").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    for j, c in enumerate(col_vals):
        cell = ws.cell(row=3, column=2 + j, value=f"{c*100:.2f}%")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, r in enumerate(row_vals):
        cell = ws.cell(row=4 + i, column=1, value=f"{r*100:.2f}%")
        cell.font = BODY_BOLD
        cell.fill = SECTION_FILL
        for j, _ in enumerate(col_vals):
            v = matrix[i][j] if i < len(matrix) and j < len(matrix[i]) else None
            ws.cell(row=4 + i, column=2 + j, value=v)
    autosize(ws)


# ============================================================
# 8. RELATIVE & SOTP
# ============================================================

def build_relative_sotp(wb, narrative, valuation):
    ws = wb.create_sheet("Relative & SOTP")
    ws["A1"] = f"{narrative['ticker']} — Relative valuation & SOTP"
    ws["A1"].font = TITLE_FONT
    currency = valuation.get("currency", "USD")

    row = 3
    # === Peer multiples ===
    section_header(ws, row, "Peer multiples", span=7)
    row += 1
    headers = ["Ticker", "NTM P/E", "EV/EBITDA", "Rev growth (TTM)", "YTD", "1Y"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    for p in narrative.get("peers", []):
        ws.cell(row=row, column=1, value=p["ticker"]).font = BODY_BOLD
        if p.get("highlight"):
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = HIGHLIGHT_FILL
        ws.cell(row=row, column=2, value=p.get("pe_ntm"))
        ws.cell(row=row, column=3, value=p.get("ev_ebitda"))
        ws.cell(row=row, column=4, value=p.get("rev_growth_ttm") / 100 if p.get("rev_growth_ttm") is not None else None).number_format = "0.0%"
        ws.cell(row=row, column=5, value=p.get("ytd") / 100 if p.get("ytd") is not None else None).number_format = "0.0%"
        ws.cell(row=row, column=6, value=p.get("y1") / 100 if p.get("y1") is not None else None).number_format = "0.0%"
        row += 1

    if narrative.get("peers_read"):
        row += 1
        ws.cell(row=row, column=1, value="Read:").font = BODY_BOLD
        ws.cell(row=row, column=2, value=narrative["peers_read"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 1

    # === Relative valuation cross-checks ===
    cc = valuation.get("cross_check")
    if cc:
        row += 1
        section_header(ws, row, "Relative valuation cross-check", span=5)
        row += 1
        for c, h in enumerate(["Method", "Metric used", "Multiple", "Implied price", "Reasoning"], start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        cc_inputs = cc.get("inputs", {})
        cc_outputs = cc.get("outputs", {})
        ws.cell(row=row, column=1, value=cc.get("name", "")).font = BODY_BOLD
        ws.cell(row=row, column=2, value=cc_inputs.get("multiple_type") or "")
        ws.cell(row=row, column=3, value=cc_inputs.get("peer_median_multiple"))
        ws.cell(row=row, column=4, value=cc_outputs.get("implied_px"))
        ws.cell(row=row, column=4).fill = HIGHLIGHT_FILL
        ws.cell(row=row, column=5, value=cc.get("reasoning", "")).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # === SOTP (only if applicable) ===
    primary = valuation.get("primary_method", {})
    if primary.get("name") == "SOTP":
        row += 1
        section_header(ws, row, "Sum-of-the-parts", span=6)
        row += 1
        for c, h in enumerate(["Segment", "Method", "Revenue (bn)", "EBITDA (bn)", "Multiple", "Segment EV (bn)"], start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for seg in primary.get("outputs", {}).get("segments", []):
            ws.cell(row=row, column=1, value=seg.get("name"))
            ws.cell(row=row, column=2, value=seg.get("method"))
            ws.cell(row=row, column=6, value=seg.get("outputs", {}).get("implied_ev_b"))
            row += 1

    autosize(ws)
    ws.column_dimensions["A"].width = 22


# ============================================================
# 9. SCENARIOS (probability-weighted)
# ============================================================

def build_scenarios(wb, valuation):
    ws = wb.create_sheet("Scenarios")
    ws["A1"] = "Scenarios + probability weights"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Final target = sum of (scenario prob × scenario implied px) + cross-check weight × cross-check implied px"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    headers = ["Scenario", "Implied px", "Probability", "Weighted contribution", "Probability reasoning", "Scenario narrative"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row = 5
    scenarios = valuation.get("scenarios", [])
    currency = valuation.get("currency", "USD")
    for scen in scenarios:
        ws.cell(row=row, column=1, value=scen["label"]).font = BODY_BOLD
        ws.cell(row=row, column=2, value=scen.get("implied_px"))
        prob = scen.get("probability")
        cell = ws.cell(row=row, column=3, value=prob)
        if prob is not None:
            cell.number_format = "0.0%"
        if prob is not None and scen.get("implied_px") is not None:
            ws.cell(row=row, column=4, value=round(prob * scen["implied_px"], 2))
        ws.cell(row=row, column=5, value=scen.get("probability_reasoning", "")).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=6, value=scen.get("reasoning", "")).alignment = Alignment(wrap_text=True, vertical="top")
        if scen["label"].lower() == "bull":
            ws.cell(row=row, column=1).fill = GOOD_FILL
        elif scen["label"].lower() == "bear":
            ws.cell(row=row, column=1).fill = BAD_FILL
        else:
            ws.cell(row=row, column=1).fill = NEUTRAL_FILL
        row += 1

    cc = valuation.get("cross_check")
    cc_weight = (valuation.get("blending_weights") or {}).get("cross_check", 0)
    if cc and cc_weight > 0:
        ws.cell(row=row, column=1, value=cc.get("name", "Cross-check")).font = BODY_BOLD
        ws.cell(row=row, column=1).fill = HIGHLIGHT_FILL
        ws.cell(row=row, column=2, value=cc.get("outputs", {}).get("implied_px"))
        ws.cell(row=row, column=3, value=cc_weight).number_format = "0.0%"
        cc_px = cc.get("outputs", {}).get("implied_px")
        if cc_px is not None:
            ws.cell(row=row, column=4, value=round(cc_weight * cc_px, 2))
        ws.cell(row=row, column=5, value=cc.get("reasoning", "")).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=6, value="Peer multiple cross-check (independent of DCF/DDM)").alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Weight total").font = BODY_BOLD
    weight_total = valuation.get("weight_total_check")
    if weight_total is None:
        weight_total = sum((s.get("probability") or 0) for s in scenarios) + cc_weight
    ws.cell(row=row, column=3, value=weight_total).number_format = "0.0%"
    if abs(weight_total - 1.0) > 0.001:
        ws.cell(row=row, column=3).fill = BAD_FILL
        ws.cell(row=row, column=5, value="WARNING: weights do not sum to 1.0").font = Font(name="Calibri", size=10, bold=True, color="991B1B")
    row += 1
    ws.cell(row=row, column=1, value="BLENDED TARGET").font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=row, column=4, value=valuation.get("blended_target")).font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=row, column=4).fill = HIGHLIGHT_FILL
    row += 1
    cp = valuation.get("current_price")
    if cp is not None:
        ws.cell(row=row, column=1, value="Current price").font = BODY_BOLD
        ws.cell(row=row, column=4, value=cp)
        upside = valuation.get("upside_pct")
        if upside is not None:
            ws.cell(row=row + 1, column=1, value="Upside / (downside) %").font = BODY_BOLD
            up_cell = ws.cell(row=row + 1, column=4, value=upside / 100)
            up_cell.number_format = "0.0%"
            up_cell.fill = GOOD_FILL if upside >= 0 else BAD_FILL

    row += 4
    cell = ws.cell(row=row, column=1, value="Why these weights?")
    cell.font = H2_FONT
    cell.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    weights_reasoning = valuation.get("weights_reasoning") or valuation.get("blending_logic", "")
    if weights_reasoning:
        c = ws.cell(row=row, column=1, value=weights_reasoning)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 60

    autosize(ws)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 55


# ============================================================
# 10. REASONING LOG
# ============================================================

def build_reasoning_log(wb, narrative, valuation):
    ws = wb.create_sheet("Reasoning log")
    ws["A1"] = "Reasoning log"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Why each method, assumption, peer set, and weight was chosen."
    ws["A2"].font = SUBTITLE_FONT

    row = 4
    primary = valuation.get("primary_method", {})

    def add(label, value):
        nonlocal row
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = BODY_BOLD
        cell.fill = SECTION_FILL
        c = ws.cell(row=row, column=2, value=value)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 2

    add("Valuation method", f"{primary.get('name')} — {primary.get('reasoning', '')}")
    method = primary.get("name")
    inputs = primary.get("inputs", {})
    if method == "DCF":
        wacc = inputs.get("wacc", {})
        add("WACC reasoning", wacc.get("reasoning", ""))
        add("Terminal growth reasoning", inputs.get("terminal_growth", {}).get("reasoning", ""))
    elif method == "DDM":
        add("DDM reasoning", inputs.get("reasoning", ""))
    elif method == "SOTP":
        for seg in primary.get("outputs", {}).get("segments", []):
            add(f"{seg['name']} ({seg['method']})", seg.get("reasoning", ""))
    cc = valuation.get("cross_check")
    if cc:
        add("Cross-check", f"{cc.get('name')} — {cc.get('reasoning', '')}")
    add("Weights reasoning", valuation.get("weights_reasoning") or valuation.get("blending_logic", ""))
    if narrative.get("data_gaps"):
        add("Data gaps", " | ".join(narrative["data_gaps"]))

    autosize(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--category", required=True, help="AI or IDX")
    parser.add_argument("--narrative", help="Path to narrative JSON")
    parser.add_argument("--valuation", help="Path to valuation JSON")
    parser.add_argument("--output", help="Path to output xlsx")
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent.parent / "output" / args.ticker
    narrative_path = Path(args.narrative) if args.narrative else (base_dir / f"{args.ticker}.json")
    valuation_path = Path(args.valuation) if args.valuation else (base_dir / f"{args.ticker}_valuation.json")
    output_path = Path(args.output) if args.output else (base_dir / f"{args.ticker}.xlsx")

    if not narrative_path.exists():
        sys.exit(f"ERROR: narrative not found: {narrative_path}")
    if not valuation_path.exists():
        sys.exit(f"ERROR: valuation not found: {valuation_path}")

    with open(narrative_path, encoding="utf-8") as f:
        narrative = json.load(f)
    with open(valuation_path, encoding="utf-8") as f:
        valuation = json.load(f)

    wb = Workbook()
    del wb["Sheet"]

    # The 10-tab institutional template
    build_cover(wb, narrative, valuation)
    build_market_stats(wb, narrative, valuation)
    build_historicals(wb, narrative)
    build_assumptions(wb, valuation)
    build_wacc_dcf(wb, valuation, narrative)
    build_formula_trace(wb, valuation)
    build_sensitivity(wb, valuation)
    build_relative_sotp(wb, narrative, valuation)
    build_scenarios(wb, valuation)
    build_reasoning_log(wb, narrative, valuation)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[OK] wrote {output_path}")


if __name__ == "__main__":
    main()
